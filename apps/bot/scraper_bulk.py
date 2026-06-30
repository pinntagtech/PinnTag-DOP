"""
Google Maps Bulk Scraper  —  Parallel Worker Pool
==================================================
Scrapes per place in one run:
  • Overview  — business info, rating summary
  • Menu & Highlights  — dish names, prices, highlight photos
  • Gallery  — all folder tabs (Food & Drinks, Vibe, Videos, …) with folder name preserved
  • Reviews  — full review data (text, rating, photos, reply thread, tags…)

Features:
  • Parallel workers  — N contexts share one Chromium instance
  • Auto-resume       — progress.json skips already-done places on re-run
  • Live dashboard    — real-time per-worker status
  • Per-place JSON    — {output_dir}/{placeId}.json
  • Dedup             — review_id deduplication built in
  • Retry-on-fail     — re-running same command retries failed places

Usage:
    python scraper_bulk.py --input places.xlsx --cookies google_cookies.json
    python scraper_bulk.py --input places.xlsx --cookies google_cookies.json --workers 6
    python scraper_bulk.py --input places.xlsx --cookies google_cookies.json --max_reviews 300
    python scraper_bulk.py --input places.xlsx --cookies google_cookies.json --skip_gallery
    python scraper_bulk.py --input places.xlsx --cookies google_cookies.json --skip_menu
"""

import argparse
import asyncio
import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

scraper_logger = logging.getLogger('pinntag-scraper')

try:
    import openpyxl
except ImportError:
    print("[ERR] pip install openpyxl"); sys.exit(1)

from playwright.async_api import async_playwright, Browser, BrowserContext, Page


# ── Face/portrait detector — DORMANT (ENABLE_FACE_FILTER=False) ──
# To enable later:
#   1. add `opencv-python-headless` + `numpy` to requirements.txt
#   2. uncomment the two imports and the two cascade loads below
#   3. set ENABLE_FACE_FILTER = True
#
# import numpy as np
# import cv2
# _FACE_FRONTAL = cv2.CascadeClassifier(
#     cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
# _FACE_PROFILE = cv2.CascadeClassifier(
#     cv2.data.haarcascades + "haarcascade_profileface.xml")
FACE_AREA_THRESHOLD = 0.12   # face box ≥ 12% of image area → portrait → drop


def _is_portrait(image_bytes: bytes) -> bool:
    """True if the image is dominated by a face (selfie/headshot/avatar).
    No-op while ENABLE_FACE_FILTER is False / opencv not imported."""
    if not ENABLE_FACE_FILTER:
        return False
    try:
        arr = np.frombuffer(image_bytes, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            return False
        h, w = img.shape[:2]
        if h == 0 or w == 0:
            return False
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        area = float(h * w)
        for cascade in (_FACE_FRONTAL, _FACE_PROFILE):
            faces = cascade.detectMultiScale(gray, scaleFactor=1.1,
                                             minNeighbors=5, minSize=(40, 40))
            for (_, _, fw, fh) in faces:
                if (fw * fh) / area >= FACE_AREA_THRESHOLD:
                    return True
            if len(faces) and 0.8 <= (w / h) <= 1.25 and max(h, w) <= 600:
                return True
        return False
    except Exception:
        return False


# ═══════════════════════════════════════════════════════════════
#  ANSI helpers
# ═══════════════════════════════════════════════════════════════

R    = "\033[0m"
BOLD = "\033[1m"
DIM  = "\033[2m"
CLR  = "\033[2K"

def _c(code, t): return f"\033[{code}m{t}{R}"
def green(t):    return _c("92", t)
def yellow(t):   return _c("93", t)
def red(t):      return _c("91", t)
def cyan(t):     return _c("96", t)
def blue(t):     return _c("94", t)
def magenta(t):  return _c("95", t)
def dim(t):      return _c("2",  t)
def bold(t):     return _c("1",  t)
def ts():        return datetime.now().strftime("%H:%M:%S")


# ═══════════════════════════════════════════════════════════════
#  Data models
# ═══════════════════════════════════════════════════════════════

@dataclass
class PlaceTask:
    place_id: str
    name: str = ""
    max_reviews: int = 200
    attempt: int = 0


@dataclass
class WorkerState:
    worker_id: int
    status: str = "idle"
    place_id: str = ""
    place_name: str = ""
    reviews_loaded: int = 0
    reviews_parsed: int = 0
    scroll_pct: float = 0.0
    current_section: str = ""   # overview | menu | gallery | reviews
    error: str = ""
    started_at: Optional[float] = None

    def elapsed(self) -> str:
        if not self.started_at: return "  0s"
        s = int(time.time() - self.started_at)
        return f"{s//60}m{s%60:02d}s" if s >= 60 else f"{s:3d}s"


# ── Review sub-models (unchanged) ────────────────────────────

@dataclass
class ReviewerProfile:
    name: str
    profile_url: Optional[str] = None
    avatar_url: Optional[str] = None
    local_guide: bool = False
    review_count: Optional[int] = None
    photo_count: Optional[int] = None

@dataclass
class OwnerReply:
    text: str
    date: Optional[str] = None
    owner_name: Optional[str] = None

@dataclass
class Review:
    review_id: Optional[str] = None
    reviewer: Optional[ReviewerProfile] = None
    rating: Optional[int] = None
    date: Optional[str] = None
    reviewed_at: Optional[str] = None
    text: Optional[str] = None
    photo_urls: list = field(default_factory=list)
    likes: Optional[int] = None
    tags: dict = field(default_factory=dict)
    price_range: Optional[str] = None
    owner_reply: Optional[OwnerReply] = None
    language: Optional[str] = None

# ── Gallery / Menu models ─────────────────────────────────────

@dataclass
class GalleryFolder:
    folder_name: str                     # "Food & drinks", "Vibe", "Videos", "All", …
    media: list = field(default_factory=list)  # list of MediaItem dicts

@dataclass
class MediaItem:
    type: str                            # "image" | "video"
    url: str                             # full-size URL
    thumbnail_url: Optional[str] = None
    caption: Optional[str] = None
    contributor: Optional[str] = None   # uploader name if available

@dataclass
class MenuItem:
    name: str
    description: Optional[str] = None
    price: Optional[str] = None
    photo_url: Optional[str] = None
    section: Optional[str] = None       # menu section heading, if any

@dataclass
class PlaceOverview:
    name: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    hours: list = field(default_factory=list)       # ["Mon: 11am–10pm", …]
    rating: Optional[float] = None
    review_count: Optional[int] = None
    price_level: Optional[str] = None              # "$", "$$", "$$$"
    category: Optional[str] = None                 # "Ramen restaurant"
    description: Optional[str] = None


# ═══════════════════════════════════════════════════════════════
#  Selectors
# ═══════════════════════════════════════════════════════════════

SEL = {
    # Reviews (unchanged)
    "review_block":      'div[data-review-id]',
    "more_button":       'button.w8nwRe',
    "review_text":       'span.wiI7pd',
    "star_img":          'span[aria-label*="star"]',
    "date":              'span.rsqaWe',
    "reviewer_name":     'div.d4r55',
    "reviewer_link":     'a.WNxzHc',
    "reviewer_avatar":   'img.NBa7we',
    "local_guide":       'div.RfnDt span',
    "likes":             'span.pkWtMe',
    "photo_buttons":     'button.Tya61d',
    "owner_reply_block": 'div.CDe7pd',
    "owner_reply_text":  'div.wiI7pd',
    "owner_reply_date":  'span.DZSIDd',
    "tag_blocks":        'div.PBK6be',
    # Overview
    "place_title":       'h1.DUwDvf, h1.fontHeadlineLarge',
    "place_category":    'button.DkEaL',
    "place_rating":      'div.F7nice span[aria-hidden="true"]',
    "place_review_count":'div.F7nice span[aria-label*="review"]',
    "place_address":     'button[data-item-id="address"] div.fontBodyMedium',
    "place_phone":       'button[data-item-id*="phone"] div.fontBodyMedium',
    "place_website":     'a[data-item-id="authority"]',
    "place_hours_btn":   'div[data-hide-tooltip-on-mouse-leave="true"] div.fontBodyMedium',
    "place_price":       'span.mgr77e',
    "place_description": 'div.PYvSYb',
    # Gallery
    "photos_tab":        'button[aria-label*="Photo"], button[jsaction*="photosTab"]',
    "gallery_folder_tabs":'div.e2moi button, button.hH0dDd, div[role="tab"]',
    "gallery_img_buttons":'button.aoRNLd, div[data-photo-index] button, button[jsaction*="pane.photo"]',
    "gallery_lightbox_img":'img.Ob1edf, img.rISBZc',
    "gallery_caption":   'div.oTLWFe, div.kSqZ2c div',
    "gallery_contributor":'div.kSqZ2c a',
    # Menu
    "menu_section":      'div.ksGge, h2.fontTitleSmall',
    "menu_item_card":    'div.Etph4b, div[jsaction*="menu.item"]',
    "menu_item_name":    'div.O4vs9b, span.fontBodyMedium',
    "menu_item_price":   'span.vlNMsc, div.dSS8se',
    "menu_item_desc":    'div.HlvMse, div.fontBodySmall',
    "menu_item_img":     'img.ITvxHb, img[src*="googleusercontent"]',
    # Highlights (overview tab cards)
    "highlight_cards":   'div.Yr7JMd, div[jsaction*="highlight"]',
    "highlight_name":    'div.qBF1Pd, span.fontBodyMedium',
    "highlight_photo":   'img.ITvxHb',
}

SCROLL_CANDIDATES = [
    'div.m6QErb[aria-label*="Reviews"]',
    'div.m6QErb[aria-label*="review"]',
    'div.m6QErb.DxyBCb',
    'div.m6QErb.WNBkOb',
    'div.m6QErb',
    'div[role="feed"]',
]

SORT_MAP = {"relevant": 0, "newest": 1, "highest": 2, "lowest": 3}


# ═══════════════════════════════════════════════════════════════
#  Pure helpers
# ═══════════════════════════════════════════════════════════════

def build_url(place_id): return f"https://www.google.com/maps/place/?q=place_id:{place_id}"

def safe_int(text):
    if not text: return None
    d = re.sub(r"[^\d]", "", str(text))
    return int(d) if d else None

def to_fullsize(url: str) -> str:
    """Strip Google's thumbnail size params to get the highest-res version."""
    if not url: return url
    url = re.sub(r'=w\d+-h\d+.*$', '=s0', url)
    url = re.sub(r'=s\d+$', '=s0', url)
    return url

def parse_relative_date(raw):
    if not raw: return None
    s = raw.strip().lower()
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    if s in ("just now", "moments ago", "a moment ago"):
        return now.strftime("%Y-%m-%dT%H:%M:%S")
    s = re.sub(r"^an?\s+", "1 ", s)
    for pattern, unit in [
        (r"(\d+)\s+second","seconds"),(r"(\d+)\s+minute","minutes"),
        (r"(\d+)\s+hour","hours"),    (r"(\d+)\s+day","days"),
        (r"(\d+)\s+week","weeks"),    (r"(\d+)\s+month","months"),
        (r"(\d+)\s+year","years"),
    ]:
        m = re.search(pattern, s)
        if m:
            n = int(m.group(1))
            d = {"seconds":timedelta(seconds=n),"minutes":timedelta(minutes=n),
                 "hours":timedelta(hours=n),"days":timedelta(days=n),
                 "weeks":timedelta(weeks=n),"months":timedelta(days=n*30),
                 "years":timedelta(days=n*365)}
            return (now - d[unit]).strftime("%Y-%m-%dT%H:%M:%S")
    return None


# ═══════════════════════════════════════════════════════════════
#  Excel / CSV reader  (unchanged)
# ═══════════════════════════════════════════════════════════════

PLACE_ID_COLS = {"place_id","placeid","place id","id","googleplaceid","google_place_id"}
NAME_COLS     = {"name","place_name","placename","business_name","title"}
URL_COLS      = {"url","link","maps_url","google_url"}

def _extract_pid_from_url(url):
    m = re.search(r"place_id[=:/]([A-Za-z0-9_\-]+)", url or "")
    return m.group(1) if m else None

def read_input_file(path: str) -> list:
    p = Path(path)
    if not p.exists():
        print(red(f"[ERR] Input file not found: {path}")); sys.exit(1)
    rows = []
    suffix = p.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1,max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i,v in enumerate(row)})
        wb.close()
    elif suffix == ".csv":
        import csv
        with open(path, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                rows.append({k.strip().lower(): str(v).strip() for k,v in row.items()})
    else:
        print(red(f"[ERR] Unsupported: {suffix}")); sys.exit(1)

    tasks = []
    for row in rows:
        pid = None
        for col,val in row.items():
            if col in PLACE_ID_COLS and val: pid = val.strip(); break
        if not pid:
            for col,val in row.items():
                if col in URL_COLS and val:
                    pid = _extract_pid_from_url(val)
                    if pid: break
        if not pid:
            for col,val in row.items():
                if isinstance(val,str) and val.startswith("ChIJ"): pid=val.strip(); break
        if not pid: continue
        name = ""
        for col,val in row.items():
            if col in NAME_COLS and val: name=val.strip(); break
        tasks.append(PlaceTask(place_id=pid, name=name))
    print(green(f"[✓] Loaded {len(tasks)} place(s) from {p.name}"))
    return tasks


# ═══════════════════════════════════════════════════════════════
#  Ledger  (resume support, unchanged)
# ═══════════════════════════════════════════════════════════════

class Ledger:
    def __init__(self, path):
        self.path  = path
        self._lock = asyncio.Lock()
        self.done: set  = set()
        self.failed: dict = {}
        self._load()

    def _load(self):
        if os.path.exists(self.path):
            try:
                d = json.loads(Path(self.path).read_text())
                self.done   = set(d.get("done",[]))
                self.failed = d.get("failed",{})
                print(cyan(f"[resume] {len(self.done)} done, {len(self.failed)} failed"))
            except Exception: pass

    async def _save(self):
        Path(self.path).write_text(json.dumps(
            {"done":sorted(self.done),"failed":self.failed,"updated_at":ts()}, indent=2))

    async def mark_done(self, pid):
        async with self._lock:
            self.done.add(pid); self.failed.pop(pid,None); await self._save()

    async def mark_failed(self, pid, err):
        async with self._lock:
            self.failed[pid] = err; await self._save()

    def is_done(self, pid): return pid in self.done


# ═══════════════════════════════════════════════════════════════
#  Dashboard  (extended with section display)
# ═══════════════════════════════════════════════════════════════

STATUS_ICONS = {
    "idle":"○", "navigating":"↗", "overview":"◉", "menu":"☰",
    "gallery":"⊞", "sorting":"⇅", "scrolling":"↓", "expanding":"⊕",
    "parsing":"◈", "saving":"⬇", "done":"✓", "error":"✗",
}
STATUS_FN = {
    "idle":dim,"navigating":cyan,"overview":magenta,"menu":yellow,
    "gallery":blue,"sorting":cyan,"scrolling":cyan,"expanding":yellow,
    "parsing":blue,"saving":green,"done":green,"error":red,
}

class Dashboard:
    def __init__(self, n, total):
        self.n=n; self.total=total; self._lock=asyncio.Lock()
        self.done=0; self.failed=0; self.start=time.time()
        self._rendered=False
        self.states=[WorkerState(i) for i in range(n)]

    def _bar(self, pct, w=20):
        f=int(pct*w); return green("█"*f)+dim("░"*(w-f))

    def _eta(self):
        el=time.time()-self.start; d=self.done
        if d==0: return "?"
        s=int(el/d*(self.total-d))
        if s<60: return f"{s}s"
        if s<3600: return f"{s//60}m{s%60:02d}s"
        return f"{s//3600}h{(s%3600)//60}m"

    def render(self):
        el  = int(time.time() - self.start)
        pct = self.done / max(self.total, 1)

        # Build lines — NO embedded \n, each entry is exactly one terminal line
        lines = []
        lines.append(bold("━" * 68))
        lines.append(
            f"  {bold('Google Maps Bulk Scraper')}  {self._bar(pct, 24)}  "
            f"{bold(str(self.done))}/{self.total}  "
            f"{dim('elapsed:')} {el//60}m{el%60:02d}s  "
            f"{dim('eta:')} {self._eta()}  "
            f"{red(f'✗{self.failed}') if self.failed else ''}"
        )
        lines.append(bold("━" * 68))

        for ws in self.states:
            fn   = STATUS_FN.get(ws.status, dim)
            icon = STATUS_ICONS.get(ws.status, "?")
            pid  = (ws.place_id[:16] + "…") if len(ws.place_id) > 17 else ws.place_id
            name = (ws.place_name[:16] + "…") if len(ws.place_name) > 17 else ws.place_name

            if ws.status == "scrolling":
                detail = f"{fn('SCROLLING')} {self._bar(ws.scroll_pct, 10)} {ws.reviews_loaded}rev"
            elif ws.status == "parsing":
                detail = f"{fn('PARSING')}   {ws.reviews_parsed} reviews"
            elif ws.status in ("overview", "menu", "gallery"):
                label  = {"overview": "OVERVIEW", "menu": "MENU", "gallery": "GALLERY"}[ws.status]
                detail = fn(label.ljust(8))
                if ws.current_section:
                    detail += f"  {dim(ws.current_section[:24])}"
            elif ws.status == "done":
                detail = green(f"DONE   {ws.reviews_parsed}rev  {ws.elapsed()}")
            elif ws.status == "error":
                detail = red(f"ERROR  {ws.error[:32]}")
            elif ws.status == "navigating":
                detail = cyan("NAVIGATING")
            elif ws.status == "saving":
                detail = green("SAVING")
            elif ws.status == "idle":
                detail = dim("IDLE")
            else:
                detail = fn(ws.status.upper().ljust(12))

            lines.append(
                f"  {dim(f'W{ws.worker_id}')}  {fn(icon)}  "
                f"{dim((name + '  ') if name else '')}"
                f"{dim(pid.ljust(18))}  {detail}  {dim(ws.elapsed())}"
            )

        lines.append(bold("━" * 68))

        if self._rendered:
            # Move cursor up exactly len(lines) rows and overwrite each line
            n = len(lines)
            print(f"\033[{n}A", end="", flush=True)
            for line in lines:
                print(f"\r{CLR}{line}", flush=True)
        else:
            # First render — just print, record line count
            for line in lines:
                print(line)
            self._rendered = True

    async def update(self, state, done_delta=0, fail_delta=0):
        async with self._lock:
            self.states[state.worker_id] = state
            self.done   += done_delta
            self.failed += fail_delta
            self.render()


# ═══════════════════════════════════════════════════════════════
#  Cookie loader  (unchanged)
# ═══════════════════════════════════════════════════════════════

def load_cookies(path):
    if not path or not os.path.exists(path): return []
    out = []
    for c in json.loads(Path(path).read_text()):
        ck = {"name":c.get("name",""),"value":c.get("value",""),
               "domain":c.get("domain",".google.com"),"path":c.get("path","/")}
        if c.get("secure") is not None: ck["secure"]=bool(c["secure"])
        if c.get("httpOnly") is not None: ck["httpOnly"]=bool(c["httpOnly"])
        if c.get("sameSite") in ("Strict","Lax","None"): ck["sameSite"]=c["sameSite"]
        exp = c.get("expires") or c.get("expirationDate")
        if exp and isinstance(exp,(int,float)) and exp>0: ck["expires"]=int(exp)
        if ck["name"] and ck["value"]: out.append(ck)
    return out


# ═══════════════════════════════════════════════════════════════
#  Scroll container finder  (unchanged)
# ═══════════════════════════════════════════════════════════════

async def _find_container(page: Page):
    best_el=None; best_score=-1
    for sel in SCROLL_CANDIDATES:
        try:
            for el in await page.locator(sel).all():
                try:
                    score = await el.evaluate(
                        "el => el.querySelectorAll('[data-review-id]').length*1000+(el.scrollHeight||0)")
                    if score > best_score: best_score=score; best_el=el
                except Exception: pass
        except Exception: pass
    return best_el


# ═══════════════════════════════════════════════════════════════
#  ① OVERVIEW scraper
# ═══════════════════════════════════════════════════════════════

async def scrape_overview(page: Page) -> dict:
    """Scrape the basic business info from the Overview tab."""
    ov = PlaceOverview()

    async def txt(sel, timeout=1500):
        try: return (await page.locator(sel).first.inner_text(timeout=timeout)).strip()
        except Exception: return None

    async def attr(sel, attribute, timeout=1500):
        try: return await page.locator(sel).first.get_attribute(attribute, timeout=timeout)
        except Exception: return None

    ov.name        = await txt(SEL["place_title"])
    ov.category    = await txt(SEL["place_category"])
    ov.address     = await txt(SEL["place_address"])
    ov.phone       = await txt(SEL["place_phone"])
    ov.website     = await attr(SEL["place_website"], "href")
    ov.description = await txt(SEL["place_description"])

    # Rating
    try:
        r = await txt(SEL["place_rating"])
        ov.rating = float(r) if r else None
    except Exception: pass

    # Review count
    try:
        rc = await txt(SEL["place_review_count"])
        ov.review_count = safe_int(rc)
    except Exception: pass

    # Price level
    try:
        pl = await txt(SEL["place_price"])
        ov.price_level = pl
    except Exception: pass

    # Hours — click the hours button to expand, then read each row
    try:
        hours_rows = await page.locator('table.eK4R0e tr').all()
        if not hours_rows:
            # Try expanding hours panel first
            hours_btn = page.locator('div[data-hide-tooltip-on-mouse-leave] button').first
            if await hours_btn.is_visible(timeout=2000):
                await hours_btn.click()
                await page.wait_for_timeout(600)
                hours_rows = await page.locator('table.eK4R0e tr').all()
        for row in hours_rows:
            try:
                day  = (await row.locator('td.ylH6lf').inner_text(timeout=500)).strip()
                time_= (await row.locator('td.mxowUb').inner_text(timeout=500)).strip()
                if day: ov.hours.append(f"{day}: {time_}")
            except Exception: pass
    except Exception: pass

    return asdict(ov)


# ═══════════════════════════════════════════════════════════════
#  ② MENU & HIGHLIGHTS scraper
# ═══════════════════════════════════════════════════════════════

async def scrape_menu(page: Page, state: WorkerState, dashboard: Dashboard) -> list:
    """
    Scrape menu highlights from the Overview tab.
    Uses JS evaluation to read directly from the DOM — no guessing class names.
    Returns list of MenuItem dicts.
    """
    items = []
    state.current_section = "menu highlights"
    await dashboard.update(state)

    # Scroll the left panel down to reveal highlight cards
    scraper_logger.info("[MENU] Scrolling to reveal highlight cards...")
    try:
        panel = page.locator('div.m6QErb').first
        for _ in range(6):
            await panel.evaluate("el => el.scrollBy(0, 350)")
            await page.wait_for_timeout(300)
    except Exception:
        pass

    # ── Extract all highlight cards via JS (avoids fragile class selectors) ──
    # Google renders highlights as buttons/divs with aria-label + an img inside
    EXTRACT_MENU_JS = r"""() => {
        const results = [];

        // Strategy 1: buttons that contain an image and aria-label (highlight cards)
        document.querySelectorAll('button[aria-label]').forEach(btn => {
            const img = btn.querySelector('img');
            const label = (btn.getAttribute('aria-label') || '').trim();
            if (!img || !label || label.length < 2) return;
            // Filter out nav/tab buttons — highlights have aria-labels that look like dish names
            if (['photos','reviews','menu','overview','search'].some(k => label.toLowerCase().includes(k))) return;

            const src = img.getAttribute('src') || '';
            if (!src.includes('googleusercontent') && !src.includes('ggpht')) return;

            // Price: look for sibling/child text matching $X.XX or $XX
            let price = null;
            const allText = btn.innerText || '';
            const priceMatch = allText.match(/[$][0-9,]+([.][0-9]{2})?/);
            if (priceMatch) price = priceMatch[0];

            results.push({
                name: label,
                price: price,
                photo_url: src.replace(/=w[0-9]+-h[0-9]+.*$/, '=s0').replace(/=s[0-9]+$/, '=s0'),
                description: null,
                section: 'highlights'
            });
        });

        // Strategy 2: divs with role=listitem inside a highlights/menu section
        document.querySelectorAll('div[role="listitem"], div[role="menuitem"]').forEach(item => {
            const nameEl = item.querySelector('div[class*="fontBody"], span[class*="fontBody"]');
            const imgEl  = item.querySelector('img');
            if (!nameEl) return;
            const name = (nameEl.innerText || '').trim();
            if (!name || name.length < 2) return;

            let price = null;
            const priceEl = item.querySelector('span[class*="price"], div[class*="price"]');
            if (priceEl) price = (priceEl.innerText || '').trim() || null;

            let photo_url = null;
            if (imgEl) {
                const src = imgEl.getAttribute('src') || '';
                if (src) photo_url = src.replace(/=w\d+-h\d+.*$/, '=s0');
            }

            results.push({ name, price, photo_url, description: null, section: 'menu' });
        });

        // Deduplicate by name
        const seen = new Set();
        return results.filter(r => {
            if (seen.has(r.name)) return false;
            seen.add(r.name);
            return true;
        });
    }"""

    try:
        raw_items = await page.evaluate(EXTRACT_MENU_JS)
        scraper_logger.info(
            f"[MENU] JS extractor found {len(raw_items or [])} items"
        )
        items = [i for i in (raw_items or []) if i.get("name")]
    except Exception:
        items = []

    # ── Try expanding "See more" / "View all" sections, then re-extract ──
    try:
        see_more_btns = await page.query_selector_all(
            'button:has-text("See more"), '
            'button:has-text("View all"), '
            'a:has-text("See more"), '
            'a:has-text("View all")'
        )
        for btn in see_more_btns[:3]:
            try:
                await btn.click()
                await page.wait_for_timeout(800)
            except Exception:
                pass

        if see_more_btns:
            try:
                expanded_items = await page.evaluate(EXTRACT_MENU_JS)
            except Exception:
                expanded_items = []

            existing_names = {
                (i.get("name") or "").lower() for i in items
            }
            for new_item in (expanded_items or []):
                nm = (new_item.get("name") or "").lower()
                if not nm or nm in existing_names:
                    continue
                items.append(new_item)
                existing_names.add(nm)
    except Exception:
        pass

    # ── Popular dishes / highlights carousel ───────────────────────────────
    try:
        carousel_items = await page.evaluate(r"""() => {
            const items = [];
            const cards = document.querySelectorAll(
                'div[data-item-id] img[src*="googleusercontent"]'
            );
            cards.forEach(img => {
                const parent = img.closest('div[data-item-id]');
                const nameEl = parent && parent.querySelector(
                    '.fontBodyMedium, .fontTitleSmall, ' +
                    'span[class*="body"], span[class*="title"]'
                );
                const name = nameEl && nameEl.textContent
                    ? nameEl.textContent.trim() : '';
                if (name && name.length >= 2) {
                    items.push({
                        name: name,
                        photo_url: img.src || '',
                        section: 'Popular',
                    });
                }
            });
            return items;
        }""")
    except Exception:
        carousel_items = []

    if carousel_items:
        existing_names = {
            (i.get("name") or "").lower() for i in items
        }
        for item in carousel_items:
            nm = (item.get("name") or "").lower()
            if not nm or nm in existing_names:
                continue
            items.append(item)
            existing_names.add(nm)

    scraper_logger.info(
        f"[MENU] After dedup: {len(items)} unique menu items"
    )
    for item in items:
        scraper_logger.debug(
            f"[MENU] Item: '{item.get('name', 'unknown')}' "
            f"price={item.get('price')} "
            f"has_photo={'yes' if item.get('photo_url') else 'no'}"
        )

    return items


# ── gallery helpers ───────────────────────────────────────────
GALLERY_SCROLL_ROUNDS = 12

# Per-business TOTAL cap (across all folders). With a per-folder cap of 15
# this lands on a folder boundary: up to 4 full folders (4×15=60), then stop
# opening more. Between-folders check only — no mid-folder interrupt needed.
MAX_GALLERY_TOTAL = int(os.getenv("MAX_GALLERY_TOTAL", "60"))

# Pseudo-folders handled by the catch-all pass — don't iterate as named folders.
SKIP_CATS = {"all", "latest", ""}

PRIORITY = ["by owner", "food & drink", "vibe", "menu",
            "exterior", "interior", "products", "videos",
            "street view & 360°", "street view"]

# Face/portrait filtering is DISABLED for now (selfies allowed through).
# Flip to True later to drop selfie/headshot/avatar photos. When enabled,
# also uncomment the opencv import + cascades below and add
# opencv-python-headless to requirements.txt.
ENABLE_FACE_FILTER = False

# Bulk URL extractor — returns {id, url, type}. id = stable photo id
# so the SAME photo at any size counts once (kills cross-folder repeats).
EXTRACT_JS = r"""() => {
  const items = [], seen = new Set();
  window.__tileReject = { notPlace: 0, avatar: 0, dup: 0, noBg: 0, kept: 0, streetview: 0 };
  const PLACE = (s) => (
    s.includes('/geo/') || s.includes('p/AF') || s.includes('geougc') ||
    s.includes('gps-cs') || s.includes('streetviewpixels') || s.includes('/p/')
  );
  const skipAvatar = (s) =>
    s.includes('/a/ACg') || s.includes('/a-/') ||
    /\/a\/[A-Za-z0-9]/.test(s) ||         // contributor profile path
    /=s(16|24|32|40|48|50|60|64)\b/.test(s);  // tiny square avatar sizes
  // NOTE: gps-cs-s/<token> place tiles have none of these → kept.
  const photoId = (u) => {
    let m;
    if (m = u.match(/gps-cs-s\/([A-Za-z0-9_\-]+)/)) return 'gps:' + m[1];
    if (m = u.match(/gps-cs[^/]*\/([A-Za-z0-9_\-]+)/)) return 'gps:' + m[1];
    if (m = u.match(/\/p\/([A-Za-z0-9_\-]+)/))       return 'p:'  + m[1];
    if (m = u.match(/geougc-cs\/([A-Za-z0-9_\-]+)/)) return 'g:'  + m[1];
    if (m = u.match(/geougc\/([A-Za-z0-9_\-]+)/))    return 'g:'  + m[1];
    if (m = u.match(/[?&]panoid=([A-Za-z0-9_\-]+)/)) return 'sv:' + m[1];
    // fallback: longest alnum token
    const toks = (u.match(/[A-Za-z0-9_\-]{20,}/g) || []);
    if (toks.length) return 'id:' + toks.sort((a,b)=>b.length-a.length)[0];
    return u.replace(/=[^/]*$/, '');
  };
  const full = (u) => {
    if (/=w\d+-h\d+/.test(u)) return u.replace(/=w\d+-h\d+[^/]*$/, '=s0');
    if (/=s\d+/.test(u))      return u.replace(/=s\d+[^/]*$/, '=s0');
    return u; // gps-cs-s URLs have no size suffix — leave as-is
  };
  const add = (src, type) => {
    if (!src) { window.__tileReject.noBg++; return; }
    // Street View / 360° panos are not business photos — drop them.
    if (type === 'image' && src.includes('streetviewpixels')) {
        window.__tileReject.streetview = (window.__tileReject.streetview || 0) + 1;
        return;
    }
    if (!src.includes('googleusercontent') && type === 'image') {
        window.__tileReject.notPlace++; return;
    }
    if (type === 'image' && (!PLACE(src) || skipAvatar(src))) {
        window.__tileReject.avatar++; return;
    }
    const id = type === 'video' ? 'v:' + src : photoId(src);
    if (seen.has(id)) { window.__tileReject.dup++; return; }
    seen.add(id);
    window.__tileReject.kept++;
    items.push({ id, url: type === 'video' ? src : full(src), type });
  };
  // The photo thumbnails live in the left rail under role="main". The
  // focused-photo overlay (role="dialog") holds only the 1 enlarged image,
  // so preferring it returned 0 tiles. Root on main where the rail is.
  const root =
    document.querySelector('div[role="main"]') || document.body;

  root.querySelectorAll('[style*="background-image"]').forEach(el => {
    const m = (el.style.backgroundImage || '').match(/url\("?([^"')]+)"?\)/);
    if (m) add(m[1], 'image');
  });
  root.querySelectorAll('img[src*="googleusercontent"], img[src*="streetviewpixels"]')
      .forEach(img => add(img.getAttribute('src') || '', 'image'));
  root.querySelectorAll('video, video source')
      .forEach(el => { const s = el.getAttribute('src') || ''; if (s.startsWith('http')) add(s, 'video'); });
  root.querySelectorAll('a[href*="youtube.com/watch"], a[href*="youtu.be/"]')
      .forEach(a => { const h = a.getAttribute('href') || ''; if (h.startsWith('http')) add(h, 'video'); });
  return items;
}"""


async def _tiles(page) -> list:
    try:
        items = await page.evaluate(EXTRACT_JS) or []
    except Exception as e:
        scraper_logger.warning(f"[TILES-DIAG] EXTRACT_JS threw: {e}")
        return []
    # one-time raw visibility: what EXTRACT_JS returned + a raw bg-div count
    try:
        raw = await page.evaluate(r"""() => {
            const root = document.querySelector('div[role="main"]') || document.body;
            const bg = [...root.querySelectorAll('[style*="background-image"]')];
            const out = [];
            for (const el of bg.slice(0, 5)) {
                const m = (el.style.backgroundImage||'').match(/url\(["']?([^"')]+)/);
                out.push(m ? m[1].slice(0,80) : '(no-url)');
            }
            return { bg_count: bg.length, sample: out };
        }""")
        scraper_logger.warning(
            f"[TILES-DIAG] EXTRACT_JS returned {len(items)} items; "
            f"raw bg-divs={raw['bg_count']}; sample={raw['sample']}")
    except Exception as e:
        scraper_logger.warning(f"[TILES-DIAG] raw probe failed: {e}")
    try:
        rej = await page.evaluate("() => window.__tileReject || null")
        scraper_logger.warning(f"[TILES-DIAG] reject counts: {rej}")
    except Exception:
        pass
    return items


async def _scroll_photos(page, rounds: int = 12):
    last = -1
    for _ in range(rounds):
        n = await page.evaluate(r"""() => {
            const root = document.querySelector('div[role="main"]') || document.body;
            // the scrollable photo container: the deepest scrollable ancestor of the tiles
            const tile = root.querySelector(
              '[style*="background-image"], img[src*="googleusercontent"], img[src*="streetviewpixels"]');
            let sc = tile;
            while (sc && sc !== document.body) {
                const oy = getComputedStyle(sc).overflowY;
                if ((oy === 'auto' || oy === 'scroll') && sc.scrollHeight > sc.clientHeight) break;
                sc = sc.parentElement;
            }
            sc = sc || root;
            sc.scrollBy(0, sc.clientHeight * 0.9);
            return root.querySelectorAll('[style*="background-image"]').length;
        }""")
        await page.wait_for_timeout(700)
        if n == last:
            break
        last = n
    return last


async def _dump_photo_dom(page, tag: str = "") -> None:
    try:
        snap = await page.evaluate(r"""() => {
            const out = { url: location.href.slice(0, 160) };
            out.dialog = !!document.querySelector('div[role="dialog"]');
            out.tabs = [...document.querySelectorAll('[role="tab"]')]
                .map(t => (t.getAttribute('aria-label') || t.innerText || '').trim())
                .filter(Boolean).slice(0, 12);
            const scrollers = [];
            document.querySelectorAll('div').forEach(d => {
                const oy = getComputedStyle(d).overflowY;
                if ((oy === 'auto' || oy === 'scroll') && d.scrollHeight > d.clientHeight + 40) {
                    scrollers.push({
                        cls: (d.className || '').toString().slice(0, 44),
                        sh: d.scrollHeight, ch: d.clientHeight,
                        tiles: d.querySelectorAll('[style*="background-image"],img[src*="googleusercontent"]').length,
                    });
                }
            });
            out.scrollers = scrollers.sort((a, b) => b.tiles - a.tiles).slice(0, 5);
            const tiles = [], seen = new Set();
            document.querySelectorAll(
                '[style*="background-image"],img[src*="googleusercontent"],img[src*="streetviewpixels"]'
            ).forEach(el => {
                if (tiles.length >= 8) return;
                let url = '';
                if (el.tagName === 'IMG') url = el.getAttribute('src') || '';
                else { const m = (el.style.backgroundImage || '').match(/url\(["']?([^"')]+)/); url = m ? m[1] : ''; }
                if (!url || seen.has(url)) return; seen.add(url);
                tiles.push({ tag: el.tagName, cls: (el.className || '').toString().slice(0, 40), url: url.slice(0, 70) });
            });
            out.tiles = tiles;
            return out;
        }""")
        scraper_logger.warning(f"[GALLERY-DOM {tag}] {snap}")
    except Exception as e:
        scraper_logger.warning(f"[GALLERY-DOM {tag}] dump failed: {e}")


async def _open_photos(page) -> bool:
    """Open the full photo viewer (not the panel preview) and confirm it by
    the presence of real folder tabs. Never the hero, never Street View."""
    # Already inside the viewer? (real folder tabs visible)
    if await _detect_categories(page):
        scraper_logger.info("[GALLERY] viewer already open")
        return True

    # Bring the "Photos & videos" section into view, then click its "All".
    try:
        await page.evaluate(r"""() => {
            const hdr = [...document.querySelectorAll('h2,h3,div')]
                .find(e => /photos?\s*&?\s*videos?/i.test(e.textContent || ''));
            if (hdr) hdr.scrollIntoView({block:'center'});
        }""")
        await page.wait_for_timeout(800)
    except Exception:
        pass

    clicked_label = await page.evaluate(r"""() => {
        const isHero = (el) => (el.getAttribute('jsaction') || '').includes('heroHeaderImage');
        const cands = [...document.querySelectorAll('button, a, [role="tab"]')];
        const el = cands.find(b => {
            if (isHero(b)) return false;
            const t = (b.getAttribute('aria-label') || b.innerText || '').trim();
            return t === 'All' || /^All\b/.test(t);
        });
        if (el) { el.click(); return (el.getAttribute('aria-label') || el.innerText || 'All').trim(); }
        return null;
    }""")
    if clicked_label:
        scraper_logger.info(f"[GALLERY] clicked entry: {clicked_label!r}")
        await page.wait_for_timeout(1800)
    else:
        scraper_logger.info("[GALLERY] no 'All' entry found to click")

    # Confirm the viewer opened: real folder tabs now present.
    if await _detect_categories(page):
        await _dump_photo_dom(page, "viewer-open")
        scraper_logger.info("[GALLERY] viewer opened (folder tabs present)")
        return True

    # Fallback: a flat gallery with tiles but no tab bar.
    if len(await _tiles(page)) >= 3:
        await _dump_photo_dom(page, "flat-open")
        scraper_logger.info("[GALLERY] flat gallery (tiles, no tabs)")
        return True

    # ── diagnostic: report what's actually on the page when we can't open ──
    try:
        diag = await page.evaluate("""() => {
            const q = (s) => document.querySelectorAll(s).length;
            return {
                tiles: q('[style*="background-image"]'),
                imgs_guc: q('img[src*="googleusercontent"]'),
                btn_photo: q('button[aria-label*="hoto"]'),
                btn_all: q('button[aria-label="All"]'),
                role_img: q('[role="img"]'),
                role_tab: q('[role="tab"]'),
                hero: q('button[jsaction*="heroHeaderImage"]'),
                consent: !!document.querySelector('form[action*="consent"], button[aria-label*="Accept"]'),
                url: location.href.slice(0, 120),
                total_nodes: document.querySelectorAll('*').length,
                body_len: (document.body ? document.body.innerText.length : 0),
                title: document.title.slice(0, 80),
            };
        }""")
    except Exception as e:
        diag = {"error": str(e)}
    scraper_logger.warning(f"[GALLERY] open failed — DOM diag: {diag}")
    return False


async def _detect_categories(page) -> list:
    """Read the photo-viewer's folder tabs generically — whatever Google
    shows for THIS business (Hairstyle, By owner, Exterior, Menu, Food &
    drink, Street View & 360°, …). Deny-list removes pseudo-folders and the
    place-panel nav tabs; everything else is a real folder. No allow-list,
    so business-specific folders are never missed."""
    try:
        cats = await page.evaluate(r"""() => {
            const out = [], seen = new Set();
            const DENY = new Set([
                'all','latest','overview','reviews','about','updates','',
                'street view & 360°','street view & 360','street view',
                'videos',
            ]);
            const push = (t) => {
                t = (t || '').trim();
                if (!t || t.length > 40) return;
                const k = t.toLowerCase();
                if (DENY.has(k) || seen.has(k)) return;
                seen.add(k); out.push(t);
            };
            document.querySelectorAll('[role="tab"]').forEach(el =>
                push(el.innerText || el.getAttribute('aria-label')));
            return out;
        }""")
    except Exception:
        cats = []
    return cats or []


async def _click_label(page, label) -> bool:
    for sel in ['[role="tab"]', 'button[aria-label]']:
        try:
            for el in await page.locator(sel).all():
                txt = ((await el.get_attribute("aria-label")) or "").strip()
                if not txt:
                    try: txt = (await el.inner_text(timeout=300)).strip()
                    except Exception: txt = ""
                if txt.lower() == label.lower():
                    await el.click(timeout=2000)
                    await page.wait_for_timeout(1200)
                    return True
        except Exception:
            continue
    return False


async def _harvest(page, folder, global_ids: set, max_items: int, label: str):
    stale = 0
    for _ in range(GALLERY_SCROLL_ROUNDS):
        added = 0
        for it in await _tiles(page):
            pid = it.get("id")
            if not pid or pid in global_ids:
                continue
            global_ids.add(pid)
            folder.media.append({
                "type": it.get("type", "image"),
                "url": it.get("url", ""),
                "thumbnail_url": it.get("url", ""),
                "caption": None, "contributor": None,
            })
            added += 1
            if len(folder.media) >= max_items:
                break
        if len(folder.media) >= max_items:
            break
        stale = stale + 1 if added == 0 else 0
        if stale >= 3:
            break
        await _scroll_photos(page)
        await page.wait_for_timeout(700)
    scraper_logger.info(f"[GALLERY] '{label}': {len(folder.media)} media")


# ═══════════════════════════════════════════════════════════════
#  GALLERY scraper — named folders + catch-all All, shared global dedup
# ═══════════════════════════════════════════════════════════════
async def scrape_gallery(page, state, dashboard,
                         max_per_folder: int = 200,
                         progress_callback=None,
                         already_open: bool = False) -> list:
    folders: list = []
    global_ids: set = set()        # shared → no repeats across any folder

    # Per-folder bound is hard-clamped to 15; the per-business total cap
    # (MAX_GALLERY_TOTAL) then stops opening new folders between folders.
    max_per_folder = min(max_per_folder or 15, 15)

    if not already_open:
        if not await _open_photos(page):
            scraper_logger.warning("[GALLERY] No photos could be opened — skipping")
            return []
    else:
        scraper_logger.info("[GALLERY] grid already open (opened by caller) — harvesting directly")
    await page.wait_for_timeout(1200)

    detected = await _detect_categories(page)
    cats = [c for c in detected if c.strip().lower() not in SKIP_CATS]
    cats = sorted(set(cats),
                  key=lambda c: PRIORITY.index(c.lower())
                  if c.lower() in PRIORITY else len(PRIORITY))
    scraper_logger.info(f"[GALLERY] Folders: {cats or '(none — flat/All only)'}")

    # named-folder pass (keeps folder names for cover-priority)
    for idx, label in enumerate(cats):
        if len(global_ids) >= MAX_GALLERY_TOTAL:
            scraper_logger.info(
                f"[GALLERY] business total cap {MAX_GALLERY_TOTAL} reached "
                f"— skipping remaining folders")
            break
        if progress_callback:
            await progress_callback("gallery", "folder_started", 0, 0,
                                    f"Scraping {label}", label)
        if not await _click_label(page, label):
            continue
        folder = GalleryFolder(folder_name=label)
        await _harvest(page, folder, global_ids, max_per_folder, label)
        if folder.media:
            folders.append(folder)
        if progress_callback:
            await progress_callback("gallery", "folder_done", idx + 1,
                                    sum(len(f.media) for f in folders),
                                    f"{label}: {len(folder.media)}", label)

    # catch-all "All" pass — picks up everything not already captured,
    # incl. businesses that only have an "All" folder. Shared global_ids
    # guarantees no duplicates vs. the named folders above. Skip entirely
    # if the business total cap was already reached by the named folders.
    if len(global_ids) >= MAX_GALLERY_TOTAL:
        scraper_logger.info(
            f"[GALLERY] business total cap reached — skipping 'All' pass")
    else:
        await _click_label(page, "All")
        await page.wait_for_timeout(800)
        leftover = GalleryFolder(folder_name="Photos")
        await _harvest(page, leftover, global_ids, max_per_folder, "Photos (All)")
        if leftover.media:
            folders.append(leftover)

    total = sum(len(f.media) for f in folders)
    scraper_logger.info(
        f"[GALLERY] DONE — {total} unique media / {len(folders)} folders: "
        f"{[(f.folder_name, len(f.media)) for f in folders]}")
    return [asdict(f) for f in folders]


# ═══════════════════════════════════════════════════════════════
#  ④ REVIEWS scraper  (unchanged logic, same function)
# ═══════════════════════════════════════════════════════════════

async def scrape_reviews(
    page: Page,
    task: PlaceTask,
    state: WorkerState,
    dashboard: Dashboard,
    sort: str,
    expand_only: bool = False,
    progress_callback=None,
) -> list:
    reviews: list[Review] = []
    seen_ids: set = set()

    async def upd(status, **kwargs):
        state.status = status
        for k,v in kwargs.items(): setattr(state,k,v)
        await dashboard.update(state)

    # Click Reviews tab
    scraper_logger.info("[REVIEWS] Clicking Reviews tab...")
    for sel in ['button[aria-label*="Reviews"]','button[aria-label*="review"]','button[data-tab-index="1"]']:
        try:
            btn = page.locator(sel).first
            if await btn.is_visible(timeout=4000):
                await btn.click(); await page.wait_for_timeout(4000)
                scraper_logger.info("[REVIEWS] Reviews tab clicked successfully")
                break
        except Exception:
            scraper_logger.warning(f"[REVIEWS] Reviews tab click failed with selector {sel}")

    try:
        await page.evaluate("""
            () => {
                const buttons = document.querySelectorAll('button');
                for (const btn of buttons) {
                    if (btn.textContent.trim().toLowerCase()
                        .includes('review')) {
                        btn.click();
                        break;
                    }
                }
            }
        """)
        await page.wait_for_timeout(3000)
    except Exception:
        pass

    # Sort
    await upd("sorting")
    sort_idx = SORT_MAP.get(sort, 1)
    if sort_idx > 0:
        for sel in ['button[aria-label*="Sort reviews"]','button[jsaction*="sortReviews"]']:
            try:
                btn = page.locator(sel).first
                if await btn.is_visible(timeout=4000):
                    await btn.click(); await page.wait_for_timeout(800)
                    items = await page.locator('li[role="menuitemradio"],div[role="menuitemradio"]').all()
                    if len(items) > sort_idx:
                        await items[sort_idx].click(); await page.wait_for_timeout(2000)
                    scraper_logger.info(f"[REVIEWS] Sort applied: {sort}")
                    break
            except Exception: pass

    # Scroll
    await upd("scrolling", reviews_loaded=0, scroll_pct=0.0)
    container = await _find_container(page)
    last_count = 0; stale = 0; target = task.max_reviews
    SCROLL_WAIT = 1500; STALE_LIMIT = 5

    if container:
        scraper_logger.info(
            f"[REVIEWS] Scroll container found — "
            f"starting scroll loop (target={target})"
        )
    else:
        scraper_logger.warning("[REVIEWS] Scroll container NOT found — using keyboard fallback")

    scroll_num = 0
    while True:
        current = await page.locator(SEL["review_block"]).count()
        scroll_num += 1
        scraper_logger.info(
            f"[SCROLL] Round #{scroll_num} — "
            f"{current}/{target} reviews in DOM | "
            f"stale={stale}/{STALE_LIMIT}"
        )
        await upd("scrolling", reviews_loaded=current, scroll_pct=min(current/max(target,1),1.0))
        if current >= target:
            scraper_logger.info(
                f"[SCROLL] Target reached: {current} reviews in DOM ✓"
            )
            break
        if current == last_count:
            stale += 1
            scraper_logger.warning(
                f"[SCROLL] No new reviews (stale tick {stale}/{STALE_LIMIT}) "
                f"— waiting longer..."
            )
            if stale >= STALE_LIMIT:
                scraper_logger.warning(
                    f"[SCROLL] Stale limit reached — "
                    f"{current} reviews loaded (max available)"
                )
                break
            await page.wait_for_timeout(SCROLL_WAIT + stale*300)
        else:
            scraper_logger.info(
                f"[SCROLL] New reviews detected: "
                f"{last_count} → {current} (+{current - last_count})"
            )
            if current != last_count and progress_callback:
                await progress_callback(
                    "reviews", "scrolling",
                    current, task.max_reviews,
                    f"Loading reviews... {current}/{task.max_reviews}",
                )
            stale=0; last_count=current
        scrolled=False
        if container:
            try:
                await container.evaluate("el=>{el.scrollTop=el.scrollHeight;el.dispatchEvent(new Event('scroll',{bubbles:true}));}")
                scraper_logger.debug("[SCROLL] scrollTop executed successfully")
                scrolled=True
            except Exception:
                container = await _find_container(page)
        if not scrolled:
            scraper_logger.warning("[SCROLL] Using keyboard End as fallback")
            await page.keyboard.press("End")
        await page.wait_for_timeout(SCROLL_WAIT)

    # Expand
    await upd("expanding")
    scraper_logger.info("[EXPAND] Looking for truncated reviews to expand...")
    total_clicked = 0
    while True:
        btns = await page.locator(f'{SEL["more_button"]}[aria-expanded="false"]').all()
        if not btns:
            scraper_logger.info("[EXPAND] All reviews expanded ✓")
            break
        scraper_logger.info(f"[EXPAND] Found {len(btns)} 'More' buttons to click")
        for i, btn in enumerate(btns):
            try:
                scraper_logger.debug(f"[EXPAND] Clicking More button #{i+1}/{len(btns)}")
                await btn.scroll_into_view_if_needed(timeout=1500)
                await btn.click(timeout=2000)
                total_clicked += 1
                await page.wait_for_timeout(100)
            except Exception: pass
        scraper_logger.info(f"[EXPAND] Expand round complete — {len(btns)} buttons clicked")
        if progress_callback:
            total_in_dom = await page.locator(SEL["review_block"]).count()
            await progress_callback(
                "reviews", "expanding",
                total_clicked, total_in_dom,
                f"Expanding {total_clicked} reviews...",
            )
        await page.wait_for_timeout(500)

    # Count how many are in DOM now
    total_in_dom = await page.locator(SEL["review_block"]).count()
    await upd("parsing", reviews_parsed=total_in_dom)

    # In hand-off mode: skip parsing, just return count
    if expand_only:
        return []   # caller will record total_in_dom separately

    # Full parse
    await upd("parsing", reviews_parsed=0)
    scraper_logger.info(
        f"[PARSE] Bulk extracting {target} reviews via JS..."
    )

    # Single JS call extracts ALL review data at once
    raw_reviews = await page.evaluate(f"""
    () => {{
        const blocks = document.querySelectorAll(
            'div[data-review-id]'
        );
        const results = [];
        const target = {target};

        for (let i = 0; i < Math.min(blocks.length, target); i++) {{
            const block = blocks[i];
            try {{
                const review = {{}};

                // Review ID
                review.review_id = block.getAttribute(
                    'data-review-id'
                );

                // Language
                let node = block;
                while (node && node !== document.body) {{
                    const l = node.getAttribute('lang');
                    if (l) {{ review.language = l; break; }}
                    node = node.parentElement;
                }}

                // Reviewer name
                const nameEl = block.querySelector('div.d4r55');
                review.reviewer_name = nameEl
                    ? nameEl.innerText.trim() : 'Unknown';

                // Reviewer link
                const linkEl = block.querySelector('a.WNxzHc');
                review.reviewer_url = linkEl
                    ? linkEl.getAttribute('href') : null;

                // Reviewer stats from aria-label
                if (linkEl) {{
                    const label = linkEl.getAttribute(
                        'aria-label'
                    ) || '';
                    const rc = label.match(/(\\d+)\\s+review/i);
                    const pc = label.match(/(\\d+)\\s+photo/i);
                    review.review_count = rc
                        ? parseInt(rc[1]) : null;
                    review.photo_count = pc
                        ? parseInt(pc[1]) : null;
                }}

                // Avatar
                const avatarEl = block.querySelector(
                    'img.NBa7we'
                );
                review.avatar_url = avatarEl
                    ? avatarEl.getAttribute('src') : null;

                // Local guide
                const guideEl = block.querySelector(
                    'div.RfnDt span'
                );
                if (guideEl) {{
                    const gt = guideEl.innerText || '';
                    review.local_guide = gt.toLowerCase()
                        .includes('local guide');
                    const rc2 = gt.match(/(\\d+)\\s+review/i);
                    const pc2 = gt.match(/(\\d+)\\s+photo/i);
                    if (rc2) review.review_count =
                        parseInt(rc2[1]);
                    if (pc2) review.photo_count =
                        parseInt(pc2[1]);
                }} else {{
                    review.local_guide = false;
                }}

                // Rating
                const starEl = block.querySelector(
                    'span[aria-label*="star"]'
                );
                if (starEl) {{
                    const m = (starEl.getAttribute(
                        'aria-label'
                    ) || '').match(/(\\d)/);
                    review.rating = m ? parseInt(m[1]) : null;
                }}

                // Date
                const dateEl = block.querySelector(
                    'span.rsqaWe'
                );
                review.date = dateEl
                    ? dateEl.innerText.trim() : null;

                // Review text
                const textEl = block.querySelector(
                    'span.wiI7pd'
                );
                review.text = textEl
                    ? textEl.innerText.trim() : null;

                // Photo URLs
                review.photo_urls = [];
                block.querySelectorAll(
                    'button.Tya61d'
                ).forEach(pb => {{
                    const style = pb.getAttribute('style') || '';
                    const m = style.match(
                        /url\\("?([^"')]+)"?\\)/
                    );
                    if (m) {{
                        const url = m[1].replace(
                            /=w\\d+-h\\d+.*$/, '=s0'
                        );
                        review.photo_urls.push(url);
                    }}
                }});

                // Likes
                const likesEl = block.querySelector(
                    'span.pkWtMe'
                );
                review.likes = likesEl
                    ? parseInt(
                        (likesEl.innerText || '0')
                        .replace(/[^\\d]/g, '')
                      ) || 0
                    : 0;

                // Tags
                review.tags = {{}};
                review.price_range = null;
                block.querySelectorAll(
                    'div.PBK6be'
                ).forEach(tb => {{
                    const text = tb.innerText.trim();
                    if (text.includes(':')) {{
                        const [k, ...v] = text.split(':');
                        review.tags[k.trim()] = v.join(':').trim();
                    }} else if (
                        text.includes('Price per person')
                    ) {{
                        const lines = text.split('\\n');
                        if (lines.length >= 2) {{
                            review.price_range =
                                lines[lines.length-1].trim();
                        }}
                    }}
                }});
                if (review.tags['Price per person']) {{
                    review.price_range =
                        review.tags['Price per person'];
                    delete review.tags['Price per person'];
                }}

                // Owner reply
                const replyBlock = block.querySelector(
                    'div.CDe7pd'
                );
                if (replyBlock) {{
                    const replyText = replyBlock.querySelector(
                        'div.wiI7pd'
                    );
                    const replyDate = replyBlock.querySelector(
                        'span.DZSIDd'
                    );
                    if (replyText) {{
                        review.owner_reply = {{
                            text: replyText.innerText.trim(),
                            date: replyDate
                                ? replyDate.innerText.trim()
                                : null,
                            owner_name: null,
                        }};
                    }}
                }}

                results.push(review);
            }} catch(e) {{
                results.push({{ error: e.toString() }});
            }}
        }}

        // Deduplicate by review_id before returning
        const seen = new Set();
        return results.filter(r => {{
            if (!r.review_id) return true;
            if (seen.has(r.review_id)) return false;
            seen.add(r.review_id);
            return true;
        }});
    }}
    """)

    scraper_logger.info(
        f"[PARSE] JS extraction complete — "
        f"{len(raw_reviews)} raw reviews (deduped in JS)"
    )

    # Convert raw JS results to Review objects
    seen_ids = set()
    errors = 0

    for i, raw in enumerate(raw_reviews):
        try:
            if raw.get('error'):
                errors += 1
                scraper_logger.warning(
                    f"[PARSE] Block #{i+1} JS error: "
                    f"{raw['error']}"
                )
                continue

            # Deduplicate
            rid = raw.get('review_id')
            if rid and rid in seen_ids:
                scraper_logger.debug(
                    f"[PARSE] Duplicate review_id skipped: "
                    f"{rid[:24]}..."
                )
                continue
            if rid:
                seen_ids.add(rid)

            # Build Review object
            review = Review()
            review.review_id = rid
            review.language = raw.get('language')
            review.rating = raw.get('rating')
            review.date = raw.get('date')
            review.reviewed_at = parse_relative_date(
                raw.get('date')
            )
            review.text = raw.get('text')
            review.photo_urls = raw.get('photo_urls', [])
            review.likes = raw.get('likes', 0)
            review.tags = raw.get('tags', {})
            review.price_range = raw.get('price_range')

            if raw.get('owner_reply'):
                review.owner_reply = OwnerReply(
                    text=raw['owner_reply']['text'],
                    date=raw['owner_reply'].get('date'),
                    owner_name=None,
                )

            review.reviewer = ReviewerProfile(
                name=raw.get('reviewer_name', 'Unknown'),
                profile_url=raw.get('reviewer_url'),
                avatar_url=raw.get('avatar_url'),
                local_guide=raw.get('local_guide', False),
                review_count=raw.get('review_count'),
                photo_count=raw.get('photo_count'),
            )

            reviews.append(review)

            if (i + 1) % 10 == 0:
                scraper_logger.info(
                    f"[PARSE] Progress: "
                    f"{len(reviews)}/{len(raw_reviews)} "
                    f"reviews processed"
                )
            if (i + 1) % 10 == 0 and progress_callback:
                await progress_callback(
                    "reviews", "parsing",
                    i + 1, len(raw_reviews),
                    f"Parsing reviews... {i+1}/{len(raw_reviews)}",
                )

        except Exception as e:
            errors += 1
            scraper_logger.warning(
                f"[PARSE] Error processing review #{i+1}: {e}"
            )

    scraper_logger.info(
        f"[PARSE] Complete — {len(reviews)} reviews parsed, "
        f"{errors} errors ✓"
    )

    await upd("parsing", reviews_parsed=len(reviews))
    return [asdict(r) for r in reviews]


async def _parse_review_block(block) -> Review:
    """Parse a single div[data-review-id] — identical to previous version."""
    review = Review()
    review.review_id = await block.get_attribute("data-review-id")

    # Language — walk up DOM
    try:
        review.language = await block.evaluate("""el => {
            let n=el; while(n&&n!==document.body){const l=n.getAttribute('lang');if(l)return l;n=n.parentElement;} return null;
        }""")
    except Exception: pass

    # Reviewer
    name="Unknown"
    try: name=(await block.locator('div.d4r55').first.inner_text(timeout=800)).strip()
    except Exception: pass

    link_href=rev_count=photo_count=avatar_url=None; local_guide=False
    try:
        link_el=block.locator('a.WNxzHc').first; link_href=await link_el.get_attribute("href")
        label=await link_el.get_attribute("aria-label") or ""
        rc=re.search(r"(\d+)\s+review",label,re.I); pc=re.search(r"(\d+)\s+photo",label,re.I)
        if rc: rev_count=int(rc.group(1))
        if pc: photo_count=int(pc.group(1))
    except Exception: pass
    try: avatar_url=await block.locator('img.NBa7we').first.get_attribute("src")
    except Exception: pass
    try:
        gt=(await block.locator('div.RfnDt span').first.inner_text(timeout=400))
        local_guide="local guide" in gt.lower()
        rc=re.search(r"(\d+)\s+review",gt,re.I); pc=re.search(r"(\d+)\s+photo",gt,re.I)
        if rc: rev_count=int(rc.group(1))
        if pc: photo_count=int(pc.group(1))
    except Exception: pass

    review.reviewer=ReviewerProfile(name=name,profile_url=link_href,avatar_url=avatar_url,
        local_guide=local_guide,review_count=rev_count,photo_count=photo_count)

    try:
        label=await block.locator('span[aria-label*="star"]').first.get_attribute("aria-label") or ""
        m=re.search(r"(\d)",label); review.rating=int(m.group(1)) if m else None
    except Exception: pass

    try:
        raw=(await block.locator('span.rsqaWe').first.inner_text(timeout=800)).strip()
        review.date=raw; review.reviewed_at=parse_relative_date(raw)
    except Exception: pass

    try: review.text=(await block.locator('span.wiI7pd').first.inner_text(timeout=800)).strip() or None
    except Exception: pass

    try:
        for pb in await block.locator('button.Tya61d').all():
            style=await pb.get_attribute("style") or ""
            m=re.search(r'url\("?([^")]+)"?\)',style)
            if m: review.photo_urls.append(to_fullsize(m.group(1)))
    except Exception: pass

    try: review.likes=safe_int(await block.locator('span.pkWtMe').first.inner_text(timeout=400))
    except Exception: review.likes=0

    try:
        for tb in await block.locator('div.PBK6be').all():
            text=(await tb.inner_text(timeout=400)).strip()
            if ":" in text:
                k,_,v=text.partition(":"); review.tags[k.strip()]=v.strip()
            elif "Price per person" in text:
                lines=text.split("\n")
                if len(lines)>=2: review.price_range=lines[-1].strip()
    except Exception: pass
    if "Price per person" in review.tags: review.price_range=review.tags.pop("Price per person")

    try:
        rb=block.locator('div.CDe7pd').first
        if await rb.count():
            rtext=(await rb.locator('div.wiI7pd').first.inner_text(timeout=800)).strip()
            rdate=None
            try: rdate=(await rb.locator('span.DZSIDd').first.inner_text(timeout=400)).strip()
            except Exception: pass
            review.owner_reply=OwnerReply(text=rtext,date=rdate)
    except Exception: pass

    return review


# ═══════════════════════════════════════════════════════════════
#  Master place scraper — orchestrates all 4 sections
# ═══════════════════════════════════════════════════════════════

async def scrape_place(
    context: BrowserContext,
    task: PlaceTask,
    state: WorkerState,
    dashboard: Dashboard,
    sort: str,
    skip_gallery: bool,
    skip_menu: bool,
    hand_off: bool = False,
) -> dict:
    """
    hand_off=False (default):  overview + menu + gallery + full review parse → save → close page
    hand_off=True:             overview + menu + gallery + scroll + expand More → save metadata
                               → Python exits but Chrome window stays open for your extension
    """
    page = await context.new_page()

    async def upd(status, **kwargs):
        state.status = status
        for k,v in kwargs.items(): setattr(state,k,v)
        await dashboard.update(state)

    result = {
        "place_id":        task.place_id,
        "place_name":      task.name,
        "scraped_at":      datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S"),
        "overview":        {},
        "menu":            [],
        "gallery":         [],
        "total_reviews":   0,
        "reviews_expanded":False,   # True once More buttons all clicked
        "reviews":         [],      # empty in hand-off mode
    }

    # Note: in hand_off mode we do NOT call page.close() — Chrome stays open
    close_page = not hand_off

    try:
        # ── Navigate ────────────────────────────────────────────
        await upd("navigating")
        await page.goto(build_url(task.place_id), wait_until="domcontentloaded", timeout=40_000)

        # Dismiss consent
        for sel in ['button[aria-label*="Accept all"]','button[aria-label*="Accept"]',
                    'button[aria-label*="Agree"]','form[action*="consent"] button']:
            try:
                btn=page.locator(sel).first
                if await btn.is_visible(timeout=1500):
                    await btn.click(); await page.wait_for_timeout(600); break
            except Exception: pass

        await page.wait_for_timeout(2000)

        # ── ① Overview ──────────────────────────────────────────
        await upd("overview", current_section="business info")
        try:
            result["overview"] = await scrape_overview(page)
            if not task.name and result["overview"].get("name"):
                task.name = result["overview"]["name"]
                state.place_name = task.name
        except Exception as e:
            result["overview"] = {"error": str(e)}

        # ── ② Menu & Highlights ─────────────────────────────────
        if not skip_menu:
            await upd("menu", current_section="highlights")
            try:
                result["menu"] = await scrape_menu(page, state, dashboard)
            except Exception:
                result["menu"] = []

        # ── ③ Gallery ───────────────────────────────────────────
        if not skip_gallery:
            await upd("gallery", current_section="opening photos tab")
            try:
                result["gallery"] = await scrape_gallery(page, state, dashboard)
            except Exception:
                result["gallery"] = []

            # Navigate back after gallery
            await page.goto(build_url(task.place_id), wait_until="domcontentloaded", timeout=30_000)
            await page.wait_for_timeout(1500)

        # ── ④ Reviews: scroll + expand (always) ─────────────────
        reviews = await scrape_reviews(
            page, task, state, dashboard, sort,
            expand_only=hand_off,   # skip parse when handing off
        )

        if hand_off:
            # Record how many are expanded in DOM, leave reviews list empty
            result["total_reviews"]    = await page.locator(SEL["review_block"]).count()
            result["reviews_expanded"] = True
            result["reviews"]          = []
        else:
            result["reviews"]        = reviews
            result["total_reviews"]  = len(reviews)

    finally:
        if close_page:
            await page.close()
        # In hand-off mode: page (and Chrome) stays open — Python will os._exit() shortly

    return result


# ═══════════════════════════════════════════════════════════════
#  Worker  (unchanged structure, passes new flags through)
# ═══════════════════════════════════════════════════════════════

async def worker(worker_id, queue, browser, cookies, ledger, dashboard,
                 output_dir, sort, skip_gallery, skip_menu, hand_off=False):
    state = WorkerState(worker_id)

    while True:
        try: task: PlaceTask = queue.get_nowait()
        except asyncio.QueueEmpty: break

        state.place_id=task.place_id; state.place_name=task.name
        state.started_at=time.time(); state.error=""

        context = await browser.new_context(
            viewport={"width":1280,"height":900}, locale="en-US",
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
        )
        if cookies: await context.add_cookies(cookies)

        try:
            data = await scrape_place(context, task, state, dashboard, sort, skip_gallery, skip_menu, hand_off=hand_off)

            state.status="saving"; await dashboard.update(state)
            out_path = output_dir / f"{task.place_id}.json"
            out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2))

            state.reviews_parsed = data["total_reviews"]
            await ledger.mark_done(task.place_id)
            await dashboard.update(state, done_delta=1)
            state.status="done"; await dashboard.update(state)

        except Exception as e:
            state.status="error"; state.error=str(e)[:80]
            await dashboard.update(state, fail_delta=1)
            await ledger.mark_failed(task.place_id, state.error)
        finally:
            if not hand_off:
                await context.close()   # normal mode: clean up
            # hand_off mode: context stays open so Chrome window survives
            queue.task_done()

        await asyncio.sleep(0.5 + worker_id * 0.1)



# ═══════════════════════════════════════════════════════════════
#  Chrome profile picker
# ═══════════════════════════════════════════════════════════════

CHROME_USER_DATA_DIRS = [
    # macOS
    Path.home() / "Library" / "Application Support" / "Google" / "Chrome",
    Path.home() / "Library" / "Application Support" / "Google" / "Chrome Canary",
    # Linux
    Path.home() / ".config" / "google-chrome",
    Path.home() / ".config" / "chromium",
    # Windows (via WSL or native)
    Path.home() / "AppData" / "Local" / "Google" / "Chrome" / "User Data",
]


def find_chrome_profiles() -> list[dict]:
    """
    Scan all known Chrome user-data directories and return every profile found.
    Each entry: { dir, profile_dir, display_name, email, user_data_dir }
    """
    profiles = []

    for user_data_dir in CHROME_USER_DATA_DIRS:
        if not user_data_dir.exists():
            continue

        # Chrome stores profiles in subdirs named Default, Profile 1, Profile 2, …
        for entry in sorted(user_data_dir.iterdir()):
            if not entry.is_dir():
                continue
            if entry.name not in ("Default",) and not entry.name.startswith("Profile"):
                continue

            prefs_file = entry / "Preferences"
            if not prefs_file.exists():
                continue

            display_name = entry.name   # fallback
            email        = ""

            try:
                prefs = json.loads(prefs_file.read_text(encoding="utf-8", errors="ignore"))
                # Real display name lives under profile.name
                pname = prefs.get("profile", {}).get("name", "")
                if pname:
                    display_name = pname
                # Account email
                acct = prefs.get("account_info", [])
                if acct and isinstance(acct, list):
                    email = acct[0].get("email", "")
                if not email:
                    email = prefs.get("profile", {}).get("last_email", "")
            except Exception:
                pass

            profiles.append({
                "user_data_dir":  str(user_data_dir),
                "profile_dir":    entry.name,
                "display_name":   display_name,
                "email":          email,
            })

    return profiles


def pick_chrome_profile() -> tuple[str, str] | None:
    """
    Interactive prompt: lists all Chrome profiles and returns
    (user_data_dir, profile_directory) for the chosen profile,
    or None if the user skips.
    """
    profiles = find_chrome_profiles()

    if not profiles:
        print(yellow("[!] No Chrome profiles found on this machine."))
        print(yellow("    Falling back to --cookies file (or anonymous session)."))
        return None

    print(bold("━" * 60))
    print(f"  {bold('Available Chrome profiles')}")
    print(bold("━" * 60))

    for i, p in enumerate(profiles, 1):
        name  = p["display_name"]
        email = f"  {dim(p['email'])}" if p["email"] else ""
        pdir  = dim(f"  [{p['profile_dir']}]")
        print(f"  {bold(cyan(str(i)))}  {name}{email}{pdir}")

    print(f"  {bold(dim('0'))}  {dim('Skip — use --cookies file or anonymous session')}")
    print(bold("━" * 60))

    while True:
        try:
            raw = input(f"  Pick a profile [0–{len(profiles)}]: ").strip()
            idx = int(raw)
            if idx == 0:
                print(dim("  → No profile selected, continuing without profile"))
                return None
            if 1 <= idx <= len(profiles):
                chosen = profiles[idx - 1]
                print(green(f"  → Using profile: {chosen['display_name']}  [{chosen['profile_dir']}]"))
                print()
                return chosen["user_data_dir"], chosen["profile_dir"]
        except (ValueError, KeyboardInterrupt):
            pass
        print(yellow(f"  Please enter a number between 0 and {len(profiles)}"))

# ═══════════════════════════════════════════════════════════════
#  Orchestrator
# ═══════════════════════════════════════════════════════════════

async def run(args):
    all_tasks = read_input_file(args.input)
    if not all_tasks: print(red("[ERR] No place IDs found")); sys.exit(1)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    ledger = Ledger(str(output_dir / "progress.json"))

    for t in all_tasks: t.max_reviews = args.max_reviews

    pending = [t for t in all_tasks if not ledger.is_done(t.place_id)]
    skipped = len(all_tasks) - len(pending)
    if skipped: print(cyan(f"[resume] Skipping {skipped} already-completed place(s)"))
    if not pending: print(green("[✓] All places done.")); return

    sections = []
    if not args.skip_menu:    sections.append("menu")
    if not args.skip_gallery: sections.append("gallery")
    sections.append("reviews")
    print(cyan(f"[→] {len(pending)} place(s)  |  {args.workers} worker(s)  |  sections: {', '.join(sections)}  |  up to {args.max_reviews} reviews"))

    cookies = load_cookies(args.cookies)
    if cookies: print(green(f"[✓] {len(cookies)} cookies loaded"))
    else: print(yellow("[!] No cookies — reviews may not load"))

    queue: asyncio.Queue = asyncio.Queue()
    for t in pending: await queue.put(t)

    dashboard = Dashboard(args.workers, len(pending))

    headless = not getattr(args, "no_headless", False)
    if not headless:
        if args.workers > 1:
            print(yellow(f"[!] --no-headless works best with --workers 1 (forcing workers=1)"))
            args.workers = 1
        print(cyan("[→] Running in visible browser mode"))

    hand_off = getattr(args, "hand_off", False)
    if hand_off:
        args.workers = 1   # hand-off is always single-place sequential
        print(cyan("[→] HAND-OFF mode: overview + menu + gallery + expand reviews → save → Chrome stays open"))

    # ── Chrome profile picker (interactive, runs before async loop) ─────────
    profile_selection = None   # (user_data_dir, profile_dir) or None

    if not headless or hand_off:
        # Only offer profile picker when a visible browser is involved
        profile_selection = pick_chrome_profile()

    # ── Detect system Chrome executable ──────────────────────────────────
    CHROME_PATHS = [
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/Applications/Google Chrome Canary.app/Contents/MacOS/Google Chrome Canary",
        "/usr/bin/google-chrome",
        "/usr/bin/google-chrome-stable",
    ]
    chrome_exe = None
    for p_path in CHROME_PATHS:
        if os.path.exists(p_path):
            chrome_exe = p_path
            break

    if chrome_exe:
        print(green(f"[✓] System Chrome: {chrome_exe}"))
    else:
        print(yellow("[!] System Chrome not found — using Playwright Chromium"))

    # ── If profile selected, cookies become optional ──────────────────────
    if profile_selection:
        if cookies:
            print(dim("[~] Profile selected — --cookies file will be ignored (profile already has session)"))
        cookies = []   # profile carries its own session; don't inject extra cookies

    async with async_playwright() as p:
        launch_args = ["--disable-blink-features=AutomationControlled"]
        if not headless:
            launch_args += ["--start-maximized"]
        if headless:
            launch_args += ["--disable-gpu", "--no-sandbox", "--disable-dev-shm-usage"]

        # ── Profile args ─────────────────────────────────────────────────
        if profile_selection:
            user_data_dir, profile_dir = profile_selection
            launch_args += [
                f"--user-data-dir={user_data_dir}",
                f"--profile-directory={profile_dir}",
            ]
            print(green(f"[✓] Launching with profile directory: {profile_dir}"))

        launch_kwargs = dict(headless=headless, args=launch_args)
        if chrome_exe:
            launch_kwargs["executable_path"] = chrome_exe

        browser = await p.chromium.launch(**launch_kwargs)
        print()
        await asyncio.gather(*[
            asyncio.create_task(worker(
                i, queue, browser, cookies, ledger, dashboard,
                output_dir, args.sort, args.skip_gallery, args.skip_menu,
                hand_off=hand_off,
            ))
            for i in range(args.workers)
        ])

        if hand_off:
            # Print handoff banner BEFORE closing Playwright context
            # We must NOT call browser.close() — Chrome window must live
            # Use os._exit() to hard-exit Python without triggering Playwright cleanup
            print(bold("═"*68))
            print(green("  ✓  Scraping complete — Chrome is yours!"))
            print(cyan("  ⬡  Browser is still open on the reviews page"))
            print(cyan("  ⬡  All More buttons have been expanded"))
            print(cyan("  ⬡  Overview + menu + gallery saved to JSON"))
            print(yellow("  ⚠  Do NOT close Chrome until your extension has finished"))
            print(f"  ⬡  Output : {output_dir.resolve()}")
            print(bold("═"*68))
            # Disconnect Playwright from the browser WITHOUT closing it
            await browser.disconnect()
        else:
            await browser.close()

    # Summary
    print(f"\n{bold('═'*68)}")
    total_done   = len(ledger.done)
    total_failed = len(ledger.failed)
    print(green(f"  ✓  Done     : {total_done}"))
    if total_failed:
        print(red(f"  ✗  Failed   : {total_failed}"))
        for pid,err in ledger.failed.items():
            print(yellow(f"       {pid[:30]}  →  {err[:50]}"))
        print(yellow("  Re-run same command to retry failed places."))
    total_reviews = sum(
        json.loads((output_dir/f"{pid}.json").read_text()).get("total_reviews",0)
        for pid in ledger.done if (output_dir/f"{pid}.json").exists()
    )
    print(cyan(f"  ⬡  Reviews  : {total_reviews} across {total_done} place(s)"))
    print(f"  ⬡  Output   : {output_dir.resolve()}")
    print(bold("═"*68))


# ═══════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Google Maps Bulk Scraper — reviews + menu + gallery",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
╔══════════════════════════════════════════════════════════════╗
║  SCRIPT GUIDE                                                ║
║  scraper_bulk.py  — bulk Excel/CSV input, parallel workers   ║
║  scraper.py       — single place_id, debug / one-off use     ║
╚══════════════════════════════════════════════════════════════╝

BULK SCRAPER (this script):
  python scraper_bulk.py --input places.xlsx --cookies google_cookies.json
  python scraper_bulk.py --input places.xlsx --cookies google_cookies.json --workers 6
  python scraper_bulk.py --input places.xlsx --cookies google_cookies.json --no-headless
  python scraper_bulk.py --input places.xlsx --cookies google_cookies.json --skip_gallery
  python scraper_bulk.py --input places.xlsx --cookies google_cookies.json --skip_menu --max_reviews 500

HAND-OFF MODE (system Chrome + profile + expand reviews + leave browser open):
  python scraper_bulk.py --input places.xlsx --hand-off
    → Script lists your Chrome profiles, you pick one, browser opens with your session + extensions
    → Scrapes overview / menu / gallery, expands all review More buttons, saves JSON, exits Python
    → Chrome stays open — run your extension on the expanded reviews page

HEADLESS BULK (no profile, parallel workers, cookies auth):
  python scraper_bulk.py --input places.xlsx --cookies google_cookies.json --workers 4

NOTE: --cookies is only needed when NOT using a Chrome profile.
      When a profile is selected, the profile's own login session is used automatically.

SINGLE PLACE (scraper.py):
  python scraper.py --place_id ChIJN1t_tDeuEmsRUsoyG83frY4 --cookies google_cookies.json
  python scraper.py --place_id ChIJN1t_tDeuEmsRUsoyG83frY4 --cookies google_cookies.json --no-headless
        """,
    )
    parser.add_argument("--input",        required=True)
    parser.add_argument("--cookies",      default=None)
    parser.add_argument("--workers",      type=int, default=4,
                        help="Parallel contexts (default 4, max 8 on local Mac)")
    parser.add_argument("--max_reviews",  type=int, default=200)
    parser.add_argument("--sort",         default="newest",
                        choices=["relevant","newest","highest","lowest"])
    parser.add_argument("--output_dir",   default="./scraped_reviews")
    parser.add_argument("--skip_gallery", action="store_true",
                        help="Skip gallery scraping (faster, reviews+menu only)")
    parser.add_argument("--skip_menu",    action="store_true",
                        help="Skip menu/highlights scraping")
    parser.add_argument("--no-headless",  action="store_true",
                        help="Show browser window (debug — auto-sets --workers 1)")
    parser.add_argument("--hand-off",     action="store_true",
                        help=(
                            "Hand-off mode: scrape overview+menu+gallery, scroll+expand reviews, "
                            "save JSON, then EXIT Python while keeping Chrome open for your extension. "
                            "Forces --workers 1 and --no-headless automatically."
                        ))
    args = parser.parse_args()

    # --hand-off implies visible browser + single worker
    if getattr(args, "hand_off", False):
        args.no_headless = True
        args.workers     = 1

    asyncio.run(run(args))

if __name__ == "__main__":
    main()
